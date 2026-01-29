"""
FastAPI 应用主入口
提供 RESTful API 和 SSE 流式接口
"""

import json
import logging
import uuid
from typing import Optional, List, Dict, Any
from datetime import datetime

from fastapi import FastAPI, HTTPException, Query, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import ValidationError
import os
import shutil
from typing import Tuple

from models import ChatRequest, Conversation, Message, get_iso_timestamp
from pydantic import BaseModel, Field
from storage import (
    save_conversation,
    load_conversation,
    list_conversations,
    delete_conversation,
    generate_conversation_title,
    generate_ai_title
)
from council import run_council
from file_storage import (
    calculate_md5,
    get_file_by_md5,
    add_file,
    get_all_files,
    delete_file as delete_file_from_storage,
    get_file_path,
    update_last_accessed
)

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# 创建 FastAPI 应用
app = FastAPI(
    title="LLM Council Simplified",
    description="简化版 LLM 委员会系统 - 四阶段协作式 AI 对话",
    version="1.0.0",
    docs_url="/docs",
    redoc_url="/redoc"
)

# CORS 配置
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


def load_config() -> Dict[str, Any]:
    """加载配置文件"""
    try:
        # 尝试多个可能的配置文件路径
        config_paths = [
            "config.json",  # 当前目录(backend/)
            "backend/config.json",  # 从项目根目录
            "../backend/config.json"  # 从其他目录
        ]
        
        for config_path in config_paths:
            try:
                with open(config_path, "r", encoding="utf-8") as f:
                    logger.info(f"成功加载配置文件: {config_path}")
                    return json.load(f)
            except FileNotFoundError:
                continue
        
        logger.error("未找到配置文件")
        return {"models": [], "chairman": "", "settings": {}}
    except Exception as e:
        logger.error(f"加载配置文件失败: {e}")
        return {"models": [], "chairman": "", "settings": {}}


def format_sse(event: str, data: dict) -> str:
    """
    格式化 SSE 事件
    
    Args:
        event: 事件名称
        data: 事件数据
    
    Returns:
        格式化的 SSE 字符串
    """
    return f"event: {event}\ndata: {json.dumps(data, ensure_ascii=False)}\n\n"


async def chat_stream_generator(request: ChatRequest, config: Dict[str, Any]):
    """
    聊天流式生成器 - 使用后台会议管理器
    
    Args:
        request: 聊天请求
        config: 配置信息
    
    Yields:
        SSE 格式的事件流
    """
    try:
        # 1. 加载或创建对话
        conv_id = request.conv_id or str(uuid.uuid4())
        conversation = load_conversation(conv_id)
        
        if conversation:
            logger.info(f"加载已有对话: {conv_id}")
        else:
            logger.info(f"创建新对话: {conv_id}")
            conversation = {
                "id": conv_id,
                "title": generate_conversation_title(request.content),
                "created_at": get_iso_timestamp(),
                "updated_at": get_iso_timestamp(),
                "messages": []
            }
        
        # 2. 检查最后一条消息是否已经是用户消息（编辑场景）
        messages = conversation.get("messages", [])
        last_message = messages[-1] if messages else None
        
        # 如果最后一条消息是用户消息且内容匹配，说明是编辑后重新生成，不需要再添加
        if last_message and last_message.get("role") == "user" and last_message.get("content") == request.content:
            logger.info(f"检测到编辑场景，使用已有的用户消息")
            user_message = last_message
        else:
            # 正常场景，添加新的用户消息
            user_message = {
                "role": "user",
                "content": request.content,
                "models": request.models,
                "attachments": [att.model_dump() for att in request.attachments] if request.attachments else [],
                "timestamp": get_iso_timestamp()
            }
            conversation["messages"].append(user_message)
            logger.info(f"添加新的用户消息")
            
            # 立即保存对话（保存用户消息）
            conversation["updated_at"] = get_iso_timestamp()
            save_conversation(conv_id, conversation)
            logger.info(f"用户消息已保存到对话: {conv_id}")
        
        # 3. 检查是否有进行中的会议
        from council_manager import council_manager
        
        existing_meetings = await council_manager.list_meetings(conv_id=conv_id)
        active_meeting = next((m for m in existing_meetings
                               if m['status'] not in ['completed', 'failed', 'cancelled']), None)
        
        if active_meeting:
            # 如果有活跃会议，重新连接它
            meeting_id = active_meeting['meeting_id']
            logger.info(f"重新连接到活跃会议: {meeting_id}")
            
            # 发送当前进度
            meeting_data = await council_manager.get_meeting(meeting_id)
            if meeting_data:
                progress = meeting_data.get('progress', {})
                
                # 发送已有的stage1结果
                if progress.get('stage1_results'):
                    for result in progress['stage1_results']:
                        yield format_sse("stage1_progress", {
                            "model": result.get("model"),
                            "status": "completed" if not result.get("error") else "error",
                            "response": result.get("response", ""),
                            "error": result.get("error")
                        })
                    yield format_sse("stage1_complete", {"results": progress['stage1_results']})
                
                # 发送已有的stage2结果
                if progress.get('stage2_results'):
                    for result in progress['stage2_results']:
                        yield format_sse("stage2_progress", result)
                    yield format_sse("stage2_complete", {"results": progress['stage2_results']})
                
                # 发送已有的stage3结果
                if progress.get('stage3_result'):
                    yield format_sse("stage3_complete", progress['stage3_result'])
                
                # 发送已有的stage4结果
                if progress.get('stage4_result'):
                    yield format_sse("stage4_complete", progress['stage4_result'])
            
            # 订阅后续更新
            queue = await council_manager.subscribe(meeting_id)
        else:
            # 创建新会议
            logger.info(f"创建新会议: 对话={conv_id}")
        
            # 准备附件数据
            attachments = []
            if request.attachments:
                attachments = [att.model_dump() for att in request.attachments]
            
            # 创建新会议
            meeting_id = await council_manager.create_meeting(
                conv_id=conv_id,
                content=request.content,
                models=request.models,
                attachments=attachments,
                config=config
            )
            logger.info(f"新会议已创建: {meeting_id}")
            
            # 发送会议ID给前端
            yield format_sse("meeting_created", {
                "meeting_id": meeting_id,
                "conv_id": conv_id
            })
            
            # 订阅会议更新
            queue = await council_manager.subscribe(meeting_id)
        
        # 4. 转发会议更新到SSE流
        try:
            while True:
                try:
                    # 等待会议更新，30秒超时发送心跳
                    update = await asyncio.wait_for(queue.get(), timeout=30.0)
                    
                    event_type = update.get("type", "update")
                    data = update.get("data", update)
                    
                    # 转发事件
                    yield format_sse(event_type, data)
                    
                    # 如果会议完成或失败，退出循环
                    if event_type in ["complete", "error"]:
                        break
                        
                except asyncio.TimeoutError:
                    # 发送心跳保持连接
                    yield format_sse("heartbeat", {"message": "keep-alive"})
                    continue
                    
        finally:
            # 取消订阅
            await council_manager.unsubscribe(meeting_id, queue)
        
        # 5. 获取最终结果并保存
        meeting_data = await council_manager.get_meeting(meeting_id)
        if meeting_data:
            progress = meeting_data.get('progress', {})
            
            stage1_results = progress.get('stage1_results', [])
            stage2_results = progress.get('stage2_results', [])
            stage3_result = progress.get('stage3_result', {})
            stage4_result = progress.get('stage4_result', {})
        
            # 6. 保存助手消息
            assistant_message = {
                "role": "assistant",
                "stage1": stage1_results,
                "stage2": stage2_results,
                "stage3": stage3_result,
                "stage4": stage4_result,
                "timestamp": get_iso_timestamp()
            }
            conversation["messages"].append(assistant_message)
            
            # 7. 如果是第一轮对话，使用AI生成标题
            if len(conversation["messages"]) == 2:  # 一条用户消息 + 一条助手消息
                try:
                    # 从config获取chairman和构建model_configs
                    chairman = config.get("chairman", "")
                    model_configs = {}
                    providers = config.get("providers", [])
                    for provider in providers:
                        provider_name = provider.get("name", "")
                        provider_models = provider.get("models", [])
                        for model in provider_models:
                            model_name = model.get("name", "")
                            full_model_name = f"{model_name}/{provider_name}"
                            model_configs[full_model_name] = {
                                "name": full_model_name,
                                "url": provider.get("url", ""),
                                "api_key": provider.get("api_key", ""),
                                "api_type": provider.get("api_type", "openai"),
                            }
                    
                    ai_title = await generate_ai_title(
                        query=request.content,
                        response=stage3_result.get("response", ""),
                        chairman_model=chairman,
                        model_configs=model_configs
                    )
                    conversation["title"] = ai_title
                    logger.info(f"AI生成对话标题: {ai_title}")
                except Exception as e:
                    logger.warning(f"AI生成标题失败，使用默认标题: {e}")
            
            # 8. 更新对话时间戳
            conversation["updated_at"] = get_iso_timestamp()
            
            # 9. 保存对话
            save_conversation(conv_id, conversation)
            
            # 10. 发送完成事件（包含更新后的标题）
            yield format_sse("complete", {
                "conv_id": conv_id,
                "title": conversation.get("title", "新对话"),
                "message": "对话完成"
            })
        
    except Exception as e:
        logger.error(f"聊天流处理错误: {e}", exc_info=True)
        yield format_sse("error", {
            "error": str(e),
            "details": "处理请求时发生错误"
        })


@app.get("/api/chat/stream")
async def chat_stream(
    conv_id: str = Query(..., description="对话 ID"),
    content: str = Query(..., description="消息内容"),
    models: List[str] = Query(..., description="模型列表"),
    attachments: Optional[str] = Query(None, description="附件 JSON 字符串")
):
    """
    发送消息并获取流式响应 (GET 方式，用于 EventSource)
    
    Args:
        conv_id: 对话 ID
        content: 消息内容
        models: 模型列表
        attachments: 附件 JSON 字符串
    
    Returns:
        SSE 流式响应
    """
    try:
        # 解析附件
        parsed_attachments = None
        if attachments:
            try:
                parsed_attachments = json.loads(attachments)
            except json.JSONDecodeError:
                logger.warning(f"无法解析附件 JSON: {attachments}")
        
        # 构建请求对象
        request = ChatRequest(
            conv_id=conv_id,
            content=content,
            models=models,
            attachments=parsed_attachments
        )
        
        # 加载配置
        config = load_config()
        
        # 验证模型是否存在
        available_models = {m["name"] for m in config.get("models", [])}
        for model in request.models:
            if model not in available_models:
                raise HTTPException(
                    status_code=400,
                    detail=f"模型 '{model}' 不存在"
                )
        
        # 返回流式响应
        return StreamingResponse(
            chat_stream_generator(request, config),
            media_type="text/event-stream",
            headers={
                "Cache-Control": "no-cache",
                "Connection": "keep-alive",
                "X-Accel-Buffering": "no"
            }
        )
        
    except ValidationError as e:
        logger.error(f"请求验证失败: {e}")
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.error(f"聊天接口错误: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Internal server error")


@app.post("/api/chat")
async def chat(request: ChatRequest):
    """
    发送消息并获取流式响应 (POST 方式，备用接口)
    
    Args:
        request: 聊天请求
    
    Returns:
        SSE 流式响应
    """
    try:
        # 加载配置
        config = load_config()
        
        # 从providers中构建可用模型列表
        available_models = set()
        providers = config.get("providers", [])
        for provider in providers:
            provider_name = provider.get("name", "")
            models = provider.get("models", [])
            for model in models:
                model_name = model.get("name", "")
                full_model_name = f"{model_name}/{provider_name}"
                available_models.add(full_model_name)
        
        # 验证模型是否存在
        for model in request.models:
            if model not in available_models:
                raise HTTPException(
                    status_code=400,
                    detail=f"模型 '{model}' 不存在"
                )
        
        # 返回流式响应
        return StreamingResponse(
            chat_stream_generator(request, config),
            media_type="text/event-stream",
            headers={
                "Cache-Control": "no-cache",
                "Connection": "keep-alive",
                "X-Accel-Buffering": "no"
            }
        )
        
    except ValidationError as e:
        logger.error(f"请求验证失败: {e}")
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.error(f"聊天接口错误: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Internal server error")


@app.get("/api/conversations")
async def get_conversations(
    limit: int = Query(50, ge=1, le=100, description="返回数量限制"),
    offset: int = Query(0, ge=0, description="偏移量"),
    sort: str = Query("created_at", pattern="^(created_at|updated_at)$", description="排序字段"),
    order: str = Query("desc", pattern="^(asc|desc)$", description="排序顺序")
):
    """
    获取对话列表
    
    Args:
        limit: 返回数量限制
        offset: 偏移量
        sort: 排序字段
        order: 排序顺序
    
    Returns:
        对话列表
    """
    try:
        # 获取所有对话
        conversations = list_conversations()
        
        # 排序
        reverse = (order == "desc")
        conversations.sort(key=lambda x: x.get(sort, ""), reverse=reverse)
        
        # 分页
        total = len(conversations)
        conversations = conversations[offset:offset + limit]
        
        return {
            "conversations": conversations,
            "total": total,
            "limit": limit,
            "offset": offset
        }
        
    except Exception as e:
        logger.error(f"获取对话列表错误: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Internal server error")


@app.post("/api/conversations/new")
async def create_conversation():
    """
    创建新对话
    
    Returns:
        新创建的对话信息
    """
    try:
        # 生成新的对话 ID
        conv_id = str(uuid.uuid4())
        
        # 创建新对话
        conversation = {
            "id": conv_id,
            "title": "新对话",
            "created_at": get_iso_timestamp(),
            "updated_at": get_iso_timestamp(),
            "messages": []
        }
        
        # 保存对话
        save_conversation(conv_id, conversation)
        
        logger.info(f"创建新对话: {conv_id}")
        
        return conversation
        
    except Exception as e:
        logger.error(f"创建对话错误: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Internal server error")


@app.get("/api/conversations/{conv_id}")
async def get_conversation(conv_id: str):
    """
    获取对话详情
    
    Args:
        conv_id: 对话 ID
    
    Returns:
        对话详情
    """
    try:
        conversation = load_conversation(conv_id)
        
        if not conversation:
            raise HTTPException(
                status_code=404,
                detail="Conversation not found"
            )
        
        return conversation
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取对话详情错误: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Internal server error")


@app.delete("/api/conversations/{conv_id}")
async def delete_conversation_endpoint(conv_id: str):
    """
    删除对话
    
    Args:
        conv_id: 对话 ID
    
    Returns:
        删除结果
    """
    try:
        success = delete_conversation(conv_id)
        
        if not success:
            raise HTTPException(
                status_code=404,
                detail="Conversation not found"
            )
        
        return {
            "message": "Conversation deleted successfully",
            "id": conv_id
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"删除对话错误: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Internal server error")


@app.get("/api/models")
async def get_models():
    """
    获取可用模型列表
    
    Returns:
        模型列表和主席模型
    """
    try:
        config = load_config()
        providers = config.get("providers", [])
        chairman = config.get("chairman", "")
        
        # 格式化模型信息 - 从所有供应商中提取模型
        formatted_models = []
        for provider in providers:
            provider_name = provider.get("name", "")
            models = provider.get("models", [])
            
            for model in models:
                model_name = model.get("name", "")
                # 模型全名格式: 模型名称/供应商
                full_model_name = f"{model_name}/{provider_name}"
                
                formatted_models.append({
                    "name": full_model_name,
                    "display_name": model.get("display_name", model_name),
                    "description": model.get("description", ""),
                    "provider": provider_name,
                    "is_chairman": full_model_name == chairman
                })
        
        return {
            "models": formatted_models,
            "chairman": chairman
        }
        
    except Exception as e:
        logger.error(f"获取模型列表错误: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Internal server error")


@app.get("/api/settings")
async def get_settings():
    """
    获取系统设置
    
    Returns:
        系统设置（温度、超时、重试次数、并发数）
    """
    try:
        config = load_config()
        settings = config.get("settings", {})
        
        return {
            "temperature": settings.get("temperature", 0.7),
            "timeout": settings.get("timeout", 120),
            "max_retries": settings.get("max_retries", 3),
            "max_concurrent": settings.get("max_concurrent", 10),
            "use_mineru": settings.get("use_mineru", False),
            "mineru_api_url": settings.get("mineru_api_url", ""),
            "mineru_api_key": settings.get("mineru_api_key", "")
        }
        
    except Exception as e:
        logger.error(f"获取设置错误: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Internal server error")


class SettingsUpdate(BaseModel):
    """设置更新请求模型"""
    temperature: float = Field(..., ge=0.0, le=2.0, description="模型温度")
    timeout: int = Field(..., ge=30, le=300, description="超时时间（秒）")
    max_retries: int = Field(..., ge=0, le=10, description="最大重试次数")
    max_concurrent: int = Field(..., ge=1, le=100, description="最大并发数")
    use_mineru: bool = Field(False, description="是否启用MinerU")
    mineru_api_url: str = Field("", description="MinerU API地址")
    mineru_api_key: str = Field("", description="MinerU API密钥")


@app.put("/api/settings")
async def update_settings(settings: SettingsUpdate):
    """
    更新系统设置
    
    Args:
        settings: 设置更新请求
    
    Returns:
        更新后的设置
    """
    temperature = settings.temperature
    timeout = settings.timeout
    max_retries = settings.max_retries
    max_concurrent = settings.max_concurrent
    try:
        # 尝试多个可能的配置文件路径
        config_paths = [
            "config.json",  # 当前目录(backend/)
            "backend/config.json",  # 从项目根目录
            "../backend/config.json"  # 从其他目录
        ]
        
        config_path = None
        for path in config_paths:
            try:
                with open(path, "r", encoding="utf-8") as f:
                    config = json.load(f)
                    config_path = path
                    break
            except FileNotFoundError:
                continue
        
        if not config_path:
            raise HTTPException(status_code=500, detail="配置文件未找到")
        
        # 更新设置
        if "settings" not in config:
            config["settings"] = {}
        
        config["settings"]["temperature"] = temperature
        config["settings"]["timeout"] = timeout
        config["settings"]["max_retries"] = max_retries
        config["settings"]["max_concurrent"] = max_concurrent
        config["settings"]["use_mineru"] = settings.use_mineru
        config["settings"]["mineru_api_url"] = settings.mineru_api_url
        config["settings"]["mineru_api_key"] = settings.mineru_api_key
        
        # 保存配置文件
        with open(config_path, "w", encoding="utf-8") as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
        
        logger.info(f"设置已更新: temperature={temperature}, timeout={timeout}, max_retries={max_retries}, max_concurrent={max_concurrent}")
        
        return {
            "temperature": temperature,
            "timeout": timeout,
            "max_retries": max_retries,
            "max_concurrent": max_concurrent,
            "use_mineru": settings.use_mineru,
            "mineru_api_url": settings.mineru_api_url,
            "mineru_api_key": settings.mineru_api_key,
            "message": "设置更新成功"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"更新设置错误: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Internal server error")


class ModelConfigUpdate(BaseModel):
    """模型配置更新请求模型"""
    models: List[Dict[str, Any]] = Field(..., description="模型配置列表")
    chairman: str = Field(..., description="主席模型名称")


@app.get("/api/models/config")
async def get_models_config():
    """
    获取完整的模型配置（包含API密钥等敏感信息）
    
    Returns:
        完整的模型配置列表和主席模型
    """
    try:
        config = load_config()
        providers = config.get("providers", [])
        chairman = config.get("chairman", "")
        
        # 构建模型配置列表（兼容旧格式）
        models = []
        for provider in providers:
            provider_name = provider.get("name", "")
            provider_models = provider.get("models", [])
            
            for model in provider_models:
                model_name = model.get("name", "")
                full_model_name = f"{model_name}/{provider_name}"
                
                models.append({
                    "name": full_model_name,
                    "display_name": model.get("display_name", model_name),
                    "description": model.get("description", ""),
                    "url": provider.get("url", ""),
                    "api_key": provider.get("api_key", ""),
                    "api_type": provider.get("api_type", "openai"),
                    "provider": provider_name
                })
        
        return {
            "models": models,
            "chairman": chairman,
            "providers": providers
        }
        
    except Exception as e:
        logger.error(f"获取模型配置错误: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Internal server error")


@app.put("/api/models/config")
async def update_models_config(config_update: ModelConfigUpdate):
    """
    更新模型配置
    
    Args:
        config_update: 模型配置更新请求
    
    Returns:
        更新后的配置
    """
    try:
        # 尝试多个可能的配置文件路径
        config_paths = [
            "config.json",  # 当前目录(backend/)
            "backend/config.json",  # 从项目根目录
            "../backend/config.json"  # 从其他目录
        ]
        
        config_path = None
        for path in config_paths:
            try:
                with open(path, "r", encoding="utf-8") as f:
                    config = json.load(f)
                    config_path = path
                    break
            except FileNotFoundError:
                continue
        
        if not config_path:
            raise HTTPException(status_code=500, detail="配置文件未找到")
        
        # 更新模型配置
        config["models"] = config_update.models
        config["chairman"] = config_update.chairman
        
        # 保存配置文件
        with open(config_path, "w", encoding="utf-8") as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
        
        logger.info(f"模型配置已更新: {len(config_update.models)} 个模型, 主席: {config_update.chairman}")
        
        return {
            "models": config_update.models,
            "chairman": config_update.chairman,
            "message": "模型配置更新成功"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"更新模型配置错误: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Internal server error")


class MessageEditRequest(BaseModel):
    """消息编辑请求模型"""
    message_index: int = Field(..., description="要编辑的消息索引")
    new_content: str = Field(..., description="新的消息内容")


@app.put("/api/conversations/{conv_id}/messages/{message_index}")
async def edit_message(conv_id: str, message_index: int, request: MessageEditRequest):
    """
    编辑对话中的消息并重新生成AI回复
    
    Args:
        conv_id: 对话ID
        message_index: 消息索引
        request: 编辑请求
        
    Returns:
        更新后的对话
    """
    try:
        # 加载对话
        conversation = load_conversation(conv_id)
        if not conversation:
            raise HTTPException(status_code=404, detail="对话不存在")
        
        messages = conversation.get("messages", [])
        
        # 验证消息索引
        if message_index < 0 or message_index >= len(messages):
            raise HTTPException(status_code=400, detail="无效的消息索引")
        
        # 验证是否为用户消息
        if messages[message_index].get("role") != "user":
            raise HTTPException(status_code=400, detail="只能编辑用户消息")
        
        # 更新消息内容和时间戳
        current_time = get_iso_timestamp()
        messages[message_index]["content"] = request.new_content
        messages[message_index]["edited"] = True
        messages[message_index]["edited_at"] = current_time
        messages[message_index]["timestamp"] = current_time  # 更新时间戳为编辑时间
        
        # 删除该消息之后的所有消息(包括对应的AI回复)
        conversation["messages"] = messages[:message_index + 1]
        
        # 更新对话时间戳
        conversation["updated_at"] = get_iso_timestamp()
        
        # 保存对话
        save_conversation(conv_id, conversation)
        
        logger.info(f"消息已编辑: 对话={conv_id}, 索引={message_index}")
        
        return conversation
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"编辑消息错误: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Internal server error")


@app.delete("/api/conversations/{conv_id}/messages/{message_index}")
async def delete_message(conv_id: str, message_index: int):
    """
    删除对话中的消息及其后续所有消息
    
    Args:
        conv_id: 对话ID
        message_index: 消息索引
        
    Returns:
        更新后的对话
    """
    try:
        # 加载对话
        conversation = load_conversation(conv_id)
        if not conversation:
            raise HTTPException(status_code=404, detail="对话不存在")
        
        messages = conversation.get("messages", [])
        
        # 验证消息索引
        if message_index < 0 or message_index >= len(messages):
            raise HTTPException(status_code=400, detail="无效的消息索引")
        
        # 验证是否为用户消息
        if messages[message_index].get("role") != "user":
            raise HTTPException(status_code=400, detail="只能删除用户消息")
        
        # 删除该消息及其后续所有消息
        conversation["messages"] = messages[:message_index]
        
        # 更新对话时间戳
        conversation["updated_at"] = get_iso_timestamp()
        
        # 保存对话
        save_conversation(conv_id, conversation)
        
        logger.info(f"消息已删除: 对话={conv_id}, 索引={message_index}")
        
        return conversation
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"删除消息错误: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Internal server error")


class ContextConfigUpdate(BaseModel):
    """上下文配置更新请求模型"""
    max_turns: int = Field(..., ge=0, le=100, description="上下文轮数")
    context_attachments: List[Dict[str, Any]] = Field(default=[], description="上下文附件列表")


@app.get("/api/conversations/{conv_id}/context")
async def get_context_config(conv_id: str):
    """
    获取对话的上下文配置
    
    Args:
        conv_id: 对话ID
        
    Returns:
        上下文配置(包含历史对话中的所有附件)
    """
    try:
        # 加载对话
        conversation = load_conversation(conv_id)
        if not conversation:
            raise HTTPException(status_code=404, detail="对话不存在")
        
        # 获取已保存的上下文配置
        context_config = conversation.get("context_config", {
            "max_turns": 3,
            "context_attachments": []
        })
        
        # 如果没有保存过上下文配置,则从历史对话中提取所有附件
        if not context_config.get("context_attachments"):
            messages = conversation.get("messages", [])
            all_attachments = []
            
            # 遍历所有用户消息,提取附件
            for msg in messages:
                if msg.get("role") == "user" and msg.get("attachments"):
                    for att in msg.get("attachments", []):
                        # 避免重复添加相同的附件
                        att_name = att.get("filename") or att.get("name")
                        if not any(a.get("filename") == att_name or a.get("name") == att_name for a in all_attachments):
                            all_attachments.append(att)
            
            context_config["context_attachments"] = all_attachments
        
        return context_config
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取上下文配置错误: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Internal server error")


@app.put("/api/conversations/{conv_id}/context")
async def update_context_config(conv_id: str, config: ContextConfigUpdate):
    """
    更新对话的上下文配置
    
    Args:
        conv_id: 对话ID
        config: 上下文配置
        
    Returns:
        更新后的配置
    """
    try:
        # 加载对话
        conversation = load_conversation(conv_id)
        if not conversation:
            raise HTTPException(status_code=404, detail="对话不存在")
        
        # 更新上下文配置
        conversation["context_config"] = {
            "max_turns": config.max_turns,
            "context_attachments": config.context_attachments
        }
        
        # 更新对话时间戳
        conversation["updated_at"] = get_iso_timestamp()
        
        # 保存对话
        save_conversation(conv_id, conversation)
        
        logger.info(f"上下文配置已更新: 对话={conv_id}, 轮数={config.max_turns}, 附件数={len(config.context_attachments)}")
        
        return conversation["context_config"]
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"更新上下文配置错误: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Internal server error")


def extract_file_content_with_mineru(file_path: str, filename: str) -> Tuple[str, Optional[str]]:
    """
    使用MinerU API提取文件内容(高质量解析)
    
    Args:
        file_path: 文件路径
        filename: 原始文件名
        
    Returns:
        (content, error) 元组
    """
    try:
        from mineru_client import MinerUClient
        import requests
        
        logger.info("=" * 60)
        logger.info("开始使用MinerU解析文档")
        logger.info(f"文件名: {filename}")
        logger.info(f"文件路径: {file_path}")
        
        # 从配置文件读取MinerU API配置
        config = load_config()
        api_key = config.get("settings", {}).get("mineru_api_key", "")
        
        if not api_key:
            logger.error("MinerU API密钥未配置")
            return "", "MinerU API密钥未配置"
        
        logger.info("MinerU API密钥已配置")
        
        # 创建MinerU客户端
        client = MinerUClient(api_token=api_key)
        logger.info("MinerU客户端创建成功")
        
        # 确定模型版本
        file_ext = os.path.splitext(filename)[1].lower()
        if file_ext == '.html':
            model_version = "MinerU-HTML"
        else:
            model_version = "pipeline"  # 默认使用pipeline模型
        
        logger.info(f"文件类型: {file_ext}, 使用模型: {model_version}")
        
        # 步骤1: 申请上传URL
        logger.info("步骤1: 申请上传URL...")
        batch_result, error = client.batch_upload_files(
            files=[{"name": filename}],
            model_version=model_version
        )
        
        if error or not batch_result:
            logger.error(f"申请上传URL失败: {error}")
            return "", f"申请上传URL失败: {error}"
        
        batch_id = batch_result.get("batch_id")
        file_urls = batch_result.get("file_urls", [])
        
        logger.info(f"获取到batch_id: {batch_id}")
        logger.info(f"获取到{len(file_urls)}个上传URL")
        
        if not file_urls:
            logger.error("未获取到上传URL")
            return "", "未获取到上传URL"
        
        upload_url = file_urls[0]
        logger.info(f"上传URL: {upload_url[:50]}...")
        
        # 步骤2: 上传文件
        logger.info("步骤2: 上传文件到MinerU...")
        error = client.upload_file_to_url(file_path, upload_url)
        if error:
            logger.error(f"文件上传失败: {error}")
            return "", f"文件上传失败: {error}"
        
        logger.info(f"文件上传成功! batch_id={batch_id}")
        logger.info("步骤3: 等待MinerU解析...")
        
        # 步骤3: 轮询查询结果
        import time
        max_wait_time = 600  # 最多等待10分钟
        poll_interval = 5  # 每5秒查询一次
        start_time = time.time()
        
        while time.time() - start_time < max_wait_time:
            if not batch_id:
                return "", "批次ID为空"
                
            results, error = client.query_batch_results(batch_id)
            
            if error:
                return "", f"查询结果失败: {error}"
            
            if results and len(results) > 0:
                result = results[0]
                state = result.get("state")
                
                if state == "done":
                    # 解析完成,下载结果
                    full_zip_url = result.get("full_zip_url")
                    if not full_zip_url:
                        logger.error("未获取到结果下载URL")
                        return "", "未获取到结果下载URL"
                    
                    logger.info(f"解析完成! 下载URL: {full_zip_url[:50]}...")
                    logger.info("步骤4: 下载并提取内容...")
                    
                    # 下载并提取内容
                    content, error = client.download_and_extract_content(full_zip_url)
                    if error:
                        logger.error(f"下载结果失败: {error}")
                        return "", f"下载结果失败: {error}"
                    
                    if content:
                        logger.info(f"MinerU解析成功! 内容长度: {len(content)} 字符")
                        logger.info(f"内容预览: {content[:200]}...")
                        logger.info("=" * 60)
                        return content, None
                    else:
                        logger.error("提取的内容为空")
                        return "", "提取的内容为空"
                    
                elif state == "failed":
                    err_msg = result.get("err_msg", "解析失败")
                    logger.error(f"MinerU解析失败: {err_msg}")
                    logger.info("=" * 60)
                    return "", f"MinerU解析失败: {err_msg}"
                    
                elif state in ["waiting-file", "pending", "running", "converting"]:
                    # 记录进度
                    progress = result.get("extract_progress", {})
                    if progress:
                        extracted = progress.get("extracted_pages", 0)
                        total = progress.get("total_pages", 0)
                        logger.info(f"MinerU解析进度: {extracted}/{total}")
                    else:
                        logger.info(f"MinerU状态: {state}")
                    
                    time.sleep(poll_interval)
                else:
                    logger.warning(f"未知状态: {state}")
                    time.sleep(poll_interval)
            else:
                time.sleep(poll_interval)
        
        logger.error(f"MinerU解析超时: 超过{max_wait_time}秒")
        logger.info("=" * 60)
        return "", f"MinerU解析超时: 超过{max_wait_time}秒"
                
    except Exception as e:
        error_msg = f"MinerU解析失败: {str(e)}"
        logger.error(error_msg, exc_info=True)
        logger.info("=" * 60)
        return "", error_msg


def extract_file_content(file_path: str, filename: str, use_mineru: bool = False) -> Tuple[str, Optional[str]]:
    """
    提取文件内容
    
    Args:
        file_path: 文件路径
        filename: 原始文件名
        use_mineru: 是否使用MinerU进行高质量解析
        
    Returns:
        (content, error) 元组，content为提取的文本内容，error为错误信息
    """
    file_ext = os.path.splitext(filename)[1].lower()
    
    # txt和md文件直接读取,不需要解析
    if file_ext in ['.txt', '.md']:
        logger.info(f"检测到文本文件({file_ext}),直接读取内容")
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
                logger.info(f"文本文件读取成功,内容长度: {len(content)} 字符")
                return content, None
        except Exception as e:
            error_msg = f"读取文本文件失败: {str(e)}"
            logger.error(error_msg)
            return "", error_msg
    
    # 其他格式:如果启用MinerU,先尝试使用MinerU解析
    if use_mineru:
        logger.info(f"检测到文档文件({file_ext}),使用MinerU解析")
        content, error = extract_file_content_with_mineru(file_path, filename)
        if content:  # MinerU解析成功
            return content, None
        # MinerU失败,继续使用本地解析
        logger.info(f"MinerU解析失败,使用本地解析: {error}")
    else:
        logger.warning(f"MinerU未启用,无法解析{file_ext}格式文件")
        return "", f"MinerU未启用,无法解析{file_ext}格式。请在设置中启用MinerU。"
    
    try:
        # Word文档 (.docx)
        if file_ext == '.docx':
            try:
                from docx import Document
                doc = Document(file_path)
                content = '\n'.join([paragraph.text for paragraph in doc.paragraphs])
                return content, None
            except ImportError:
                return "", "需要安装python-docx库来处理.docx文件"
            except Exception as e:
                return "", f"读取.docx文件失败: {str(e)}"
        
        # Excel文件 (.xlsx, .xls)
        elif file_ext in ['.xlsx', '.xls']:
            try:
                from openpyxl import load_workbook
                wb = load_workbook(file_path, data_only=True)
                content_parts = []
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    content_parts.append(f"=== 工作表: {sheet_name} ===")
                    for row in sheet.iter_rows(values_only=True):
                        row_text = '\t'.join([str(cell) if cell is not None else '' for cell in row])
                        if row_text.strip():
                            content_parts.append(row_text)
                return '\n'.join(content_parts), None
            except ImportError:
                return "", "需要安装openpyxl库来处理Excel文件"
            except Exception as e:
                return "", f"读取Excel文件失败: {str(e)}"
        
        # PDF文件
        elif file_ext == '.pdf':
            try:
                from PyPDF2 import PdfReader
                reader = PdfReader(file_path)
                content_parts = []
                for i, page in enumerate(reader.pages):
                    text = page.extract_text()
                    if text.strip():
                        content_parts.append(f"=== 第{i+1}页 ===\n{text}")
                return '\n\n'.join(content_parts), None
            except ImportError:
                return "", "需要安装PyPDF2库来处理PDF文件"
            except Exception as e:
                return "", f"读取PDF文件失败: {str(e)}"
        
        # DOC文件 (旧版Word)
        elif file_ext == '.doc':
            return "", ".doc格式不支持，请转换为.docx格式"
        
        else:
            return "", f"不支持的文件格式: {file_ext}"
            
    except Exception as e:
        logger.error(f"提取文件内容错误: {e}", exc_info=True)
        return "", f"提取文件内容失败: {str(e)}"


@app.post("/api/upload")
async def upload_file(file: UploadFile = File(...)):
    """
    上传文件并提取内容(支持MD5去重)
    
    支持的文件格式:
    - .txt: 纯文本文件
    - .docx: Word文档
    - .xlsx, .xls: Excel表格
    - .pdf: PDF文档
    
    Args:
        file: 上传的文件
        
    Returns:
        上传结果，包含文件名、路径、大小、MD5和提取的文本内容
    """
    try:
        logger.info("=" * 60)
        logger.info("收到文件上传请求")
        logger.info(f"文件名: {file.filename}")
        logger.info(f"Content-Type: {file.content_type}")
        
        # 从配置文件读取是否启用MinerU
        config = load_config()
        use_mineru = config.get("settings", {}).get("use_mineru", False)
        
        logger.info(f"MinerU状态: {'启用' if use_mineru else '禁用'}")
        
        # 检查文件类型
        file_ext = os.path.splitext(file.filename)[1].lower()
        
        # 根据MinerU状态确定允许的文件类型
        if use_mineru:
            # 启用MinerU时,支持多种文档格式
            allowed_extensions = ['.txt', '.md', '.doc', '.docx', '.xlsx', '.xls', '.pdf', '.ppt', '.pptx', '.png', '.jpg', '.jpeg', '.html']
            logger.info("MinerU已启用,支持多种文档格式")
        else:
            # 未启用MinerU时,只支持txt和markdown
            allowed_extensions = ['.txt', '.md']
            logger.info("MinerU未启用,仅支持txt和markdown文件")
        
        logger.info(f"文件扩展名: {file_ext}")
        logger.info(f"允许的格式: {', '.join(allowed_extensions)}")
        
        if file_ext not in allowed_extensions:
            logger.error(f"不支持的文件格式: {file_ext}")
            if not use_mineru:
                raise HTTPException(
                    status_code=400,
                    detail=f"未启用MinerU,仅支持txt和markdown文件。如需上传其他格式,请在设置中启用MinerU。"
                )
            else:
                raise HTTPException(
                    status_code=400,
                    detail=f"不支持的文件格式。支持的格式: {', '.join(allowed_extensions)}"
                )
        
        # 创建uploads目录（如果不存在）
        upload_dir = "backend/uploads"
        os.makedirs(upload_dir, exist_ok=True)
        
        # 生成临时文件名
        temp_filename = f"temp_{uuid.uuid4()}{file_ext}"
        temp_path = os.path.join(upload_dir, temp_filename)
        
        # 保存临时文件
        with open(temp_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        
        # 计算MD5
        file_md5 = calculate_md5(temp_path)
        logger.info(f"文件MD5: {file_md5}")
        
        # 检查是否已存在相同MD5的文件
        existing_file = get_file_by_md5(file_md5)
        
        if existing_file:
            # 文件已存在,删除临时文件,返回已有文件信息
            os.remove(temp_path)
            logger.info(f"文件已存在(MD5: {file_md5}),复用已有文件")
            
            # 更新最后访问时间
            update_last_accessed(file_md5)
            
            return {
                "filename": file.filename,
                "name": file.filename,
                "path": existing_file["stored_path"],
                "size": existing_file["size"],
                "content_type": file.content_type,
                "content": existing_file["content"],
                "content_length": len(existing_file["content"]),
                "md5": file_md5,
                "is_duplicate": True,
                "extraction_error": None
            }
        
        # 新文件,重命名为正式文件名
        unique_filename = f"{uuid.uuid4()}{file_ext}"
        file_path = os.path.join(upload_dir, unique_filename)
        os.rename(temp_path, file_path)
        
        # 提取文件内容
        content, error = extract_file_content(file_path, file.filename, use_mineru=use_mineru)
        
        if error:
            logger.warning(f"文件内容提取失败: {error}")
        
        # 添加到文件存储系统
        file_info = add_file(
            file_path=file_path,
            original_filename=file.filename,
            content=content,
            size=os.path.getsize(file_path),
            md5=file_md5
        )
        
        logger.info(f"文件上传成功: {file.filename} -> {file_path}, MD5: {file_md5}, 内容长度: {len(content)}")
        
        return {
            "filename": file.filename,
            "name": file.filename,
            "path": file_path,
            "size": file_info["size"],
            "content_type": file.content_type,
            "content": content,
            "content_length": len(content),
            "md5": file_md5,
            "is_duplicate": False,
            "extraction_error": error
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"文件上传错误: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"文件上传失败: {str(e)}")


@app.get("/api/files")
async def list_files_endpoint():
    """
    获取所有已上传文件列表
    
    Returns:
        文件列表
    """
    try:
        files = get_all_files()
        return {
            "files": files,
            "total": len(files)
        }
    except Exception as e:
        logger.error(f"获取文件列表错误: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Internal server error")


@app.get("/api/files/{md5}/download")
async def download_file(md5: str):
    """
    下载文件(源文件)
    
    Args:
        md5: 文件MD5值
        
    Returns:
        文件内容
    """
    try:
        from fastapi.responses import FileResponse
        
        file_path = get_file_path(md5)
        if not file_path or not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="文件不存在")
        
        # 更新最后访问时间
        update_last_accessed(md5)
        
        # 获取文件信息
        file_info = get_file_by_md5(md5)
        filename = file_info.get("filename", "download")
        
        return FileResponse(
            path=file_path,
            filename=filename,
            media_type='application/octet-stream'
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"下载文件错误: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Internal server error")


@app.delete("/api/files/{md5}")
async def delete_file_endpoint(md5: str):
    """
    删除文件
    
    Args:
        md5: 文件MD5值
        
    Returns:
        删除结果
    """
    try:
        success = delete_file_from_storage(md5)
        
        if not success:
            raise HTTPException(status_code=404, detail="文件不存在")
        
        logger.info(f"文件已删除: MD5={md5}")
        
        return {
            "message": "文件删除成功",
            "md5": md5
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"删除文件错误: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Internal server error")


# ==================== 供应商管理 API ====================

from provider_manager import (
    load_providers,
    add_provider,
    update_provider,
    delete_provider,
    fetch_provider_models,
    test_model,
    get_provider_models,
    add_model_to_provider,
    delete_model_from_provider
)


class ProviderCreate(BaseModel):
    """创建供应商请求模型"""
    name: str = Field(..., description="供应商名称")
    url: str = Field(..., description="API URL")
    api_key: str = Field(..., description="API密钥")
    api_type: str = Field(..., description="API类型 (openai/anthropic)")


class ProviderUpdate(BaseModel):
    """更新供应商请求模型"""
    url: Optional[str] = Field(None, description="API URL")
    api_key: Optional[str] = Field(None, description="API密钥")
    api_type: Optional[str] = Field(None, description="API类型 (openai/anthropic)")


class ModelAddRequest(BaseModel):
    """添加模型到本地配置请求模型"""
    provider_name: str = Field(..., description="供应商名称")
    model_id: str = Field(..., description="模型ID")
    display_name: str = Field(..., description="显示名称")
    description: Optional[str] = Field("", description="模型描述")


@app.get("/api/providers")
async def get_providers():
    """
    获取所有供应商列表
    
    Returns:
        供应商列表
    """
    try:
        providers = load_providers()
        # 隐藏API密钥
        for provider in providers:
            if "api_key" in provider:
                provider["api_key_masked"] = "*" * 20
                del provider["api_key"]
        
        return {
            "providers": providers,
            "total": len(providers)
        }
    except Exception as e:
        logger.error(f"获取供应商列表错误: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Internal server error")


@app.post("/api/providers")
async def create_provider(provider: ProviderCreate):
    """
    添加新供应商
    
    Args:
        provider: 供应商信息
    
    Returns:
        创建结果
    """
    try:
        success, error = add_provider(
            name=provider.name,
            url=provider.url,
            api_key=provider.api_key,
            api_type=provider.api_type
        )
        
        if not success:
            raise HTTPException(status_code=400, detail=error)
        
        logger.info(f"供应商已添加: {provider.name}")
        
        return {
            "message": "供应商添加成功",
            "name": provider.name
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"添加供应商错误: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Internal server error")


@app.put("/api/providers/{provider_name}")
async def update_provider_endpoint(provider_name: str, provider: ProviderUpdate):
    """
    更新供应商配置
    
    Args:
        provider_name: 供应商名称
        provider: 更新的供应商信息
    
    Returns:
        更新结果
    """
    try:
        success, error = update_provider(
            name=provider_name,
            url=provider.url,
            api_key=provider.api_key,
            api_type=provider.api_type
        )
        
        if not success:
            raise HTTPException(status_code=400, detail=error)
        
        logger.info(f"供应商已更新: {provider_name}")
        
        return {
            "message": "供应商更新成功",
            "name": provider_name
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"更新供应商错误: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Internal server error")


@app.delete("/api/providers/{provider_name}")
async def delete_provider_endpoint(provider_name: str):
    """
    删除供应商
    
    Args:
        provider_name: 供应商名称
    
    Returns:
        删除结果
    """
    try:
        success, error = delete_provider(provider_name)
        
        if not success:
            raise HTTPException(status_code=400, detail=error)
        
        logger.info(f"供应商已删除: {provider_name}")
        
        return {
            "message": "供应商删除成功",
            "name": provider_name
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"删除供应商错误: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Internal server error")


@app.get("/api/providers/{provider_name}/models")
async def get_provider_models_endpoint(provider_name: str):
    """
    获取供应商下已添加的模型列表
    
    Args:
        provider_name: 供应商名称
    
    Returns:
        模型列表
    """
    try:
        models, error = get_provider_models(provider_name)
        
        if error:
            raise HTTPException(status_code=400, detail=error)
        
        return {
            "provider": provider_name,
            "models": models,
            "total": len(models) if models else 0
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取供应商模型列表错误: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Internal server error")


@app.get("/api/providers/{provider_name}/models/fetch")
async def fetch_provider_models_endpoint(provider_name: str):
    """
    从供应商API获取可用模型列表
    
    Args:
        provider_name: 供应商名称
    
    Returns:
        模型列表
    """
    try:
        models, error = await fetch_provider_models(provider_name)
        
        if error:
            raise HTTPException(status_code=400, detail=error)
        
        return {
            "provider": provider_name,
            "models": models,
            "total": len(models) if models else 0
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取供应商API模型列表错误: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Internal server error")


@app.post("/api/providers/models/test")
async def test_provider_model(provider_name: str = Query(...), model_name: str = Query(...)):
    """
    测试模型是否可用
    
    Args:
        provider_name: 供应商名称
        model_name: 模型名称
    
    Returns:
        测试结果
    """
    try:
        success, response, error = await test_model(provider_name, model_name)
        
        return {
            "provider": provider_name,
            "model": model_name,
            "success": success,
            "response": response,
            "error": error
        }
    except Exception as e:
        logger.error(f"测试模型错误: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Internal server error")


@app.post("/api/providers/{provider_name}/models")
async def add_model_to_provider_endpoint(provider_name: str, request: ModelAddRequest):
    """
    添加模型到供应商
    
    Args:
        provider_name: 供应商名称
        request: 添加模型请求
    
    Returns:
        添加结果
    """
    try:
        success, error = add_model_to_provider(
            provider_name=provider_name,
            model_name=request.model_id,
            display_name=request.display_name,
            description=request.description
        )
        
        if not success:
            raise HTTPException(status_code=400, detail=error)
        
        logger.info(f"模型已添加: {request.model_id} (供应商: {provider_name})")
        
        return {
            "message": "模型添加成功",
            "provider": provider_name,
            "model_name": request.model_id
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"添加模型错误: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Internal server error")


@app.delete("/api/providers/{provider_name}/models/{model_name}")
async def delete_model_from_provider_endpoint(provider_name: str, model_name: str):
    """
    从供应商删除模型
    
    Args:
        provider_name: 供应商名称
        model_name: 模型名称
    
    Returns:
        删除结果
    """
    try:
        success, error = delete_model_from_provider(provider_name, model_name)
        
        if not success:
            raise HTTPException(status_code=400, detail=error)
        
        logger.info(f"模型已删除: {model_name} (供应商: {provider_name})")
        
        return {
            "message": "模型删除成功",
            "provider": provider_name,
            "model_name": model_name
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"删除模型错误: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Internal server error")


# ==================== 会议管理 API ====================

from council_manager import council_manager
import asyncio


@app.post("/api/meetings/start")
async def start_meeting(request: ChatRequest):
    """
    启动新会议（后台运行）
    
    Args:
        request: 聊天请求
        
    Returns:
        会议ID和初始状态
    """
    try:
        # 加载配置
        config = load_config()
        
        # 从providers中构建可用模型列表
        available_models = set()
        providers = config.get("providers", [])
        for provider in providers:
            provider_name = provider.get("name", "")
            models = provider.get("models", [])
            for model in models:
                model_name = model.get("name", "")
                full_model_name = f"{model_name}/{provider_name}"
                available_models.add(full_model_name)
        
        # 验证模型是否存在
        for model in request.models:
            if model not in available_models:
                raise HTTPException(
                    status_code=400,
                    detail=f"模型 '{model}' 不存在"
                )
        
        # 准备附件数据
        attachments = []
        if request.attachments:
            attachments = [att.model_dump() for att in request.attachments]
        
        # 创建会议
        meeting_id = await council_manager.create_meeting(
            conv_id=request.conv_id,
            content=request.content,
            models=request.models,
            attachments=attachments,
            config=config
        )
        
        logger.info(f"会议已启动: {meeting_id}, 对话: {request.conv_id}")
        
        return {
            "meeting_id": meeting_id,
            "conv_id": request.conv_id,
            "message": "会议已在后台启动"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"启动会议错误: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/meetings/{meeting_id}")
async def get_meeting_status(meeting_id: str):
    """
    获取会议状态
    
    Args:
        meeting_id: 会议ID
        
    Returns:
        会议状态和进度
    """
    try:
        meeting = await council_manager.get_meeting(meeting_id)
        
        if not meeting:
            raise HTTPException(status_code=404, detail="会议不存在")
        
        return meeting
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取会议状态错误: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Internal server error")


@app.get("/api/meetings/{meeting_id}/stream")
async def stream_meeting_updates(meeting_id: str):
    """
    订阅会议更新流（SSE）
    
    Args:
        meeting_id: 会议ID
        
    Returns:
        SSE 流式响应
    """
    async def event_generator():
        try:
            # 先获取会议当前状态，发送历史进度
            meeting_data = await council_manager.get_meeting(meeting_id)
            if meeting_data:
                progress = meeting_data.get('progress', {})
                
                # 发送已有的stage1结果
                if progress.get('stage1_results'):
                    for result in progress['stage1_results']:
                        yield format_sse("stage1_progress", {
                            "model": result.get("model"),
                            "status": "completed" if not result.get("error") else "error",
                            "response": result.get("response", ""),
                            "error": result.get("error")
                        })
                    yield format_sse("stage1_complete", {"results": progress['stage1_results']})
                
                # 发送已有的stage2结果
                if progress.get('stage2_results'):
                    for result in progress['stage2_results']:
                        yield format_sse("stage2_progress", result)
                    yield format_sse("stage2_complete", {"results": progress['stage2_results']})
                
                # 发送已有的stage3结果
                if progress.get('stage3_result'):
                    yield format_sse("stage3_complete", progress['stage3_result'])
                
                # 发送已有的stage4结果
                if progress.get('stage4_result'):
                    yield format_sse("stage4_complete", progress['stage4_result'])
            
            # 订阅会议更新
            queue = await council_manager.subscribe(meeting_id)
            
            try:
                while True:
                    # 从队列获取更新
                    try:
                        update = await asyncio.wait_for(queue.get(), timeout=30.0)
                        
                        # 根据更新类型发送不同的事件
                        event_type = update.get("type", "update")
                        yield format_sse(event_type, update.get("data", update))
                        
                        # 如果会议完成或失败，结束流
                        if event_type in ["complete", "error"]:
                            break
                            
                    except asyncio.TimeoutError:
                        # 发送心跳保持连接
                        yield format_sse("heartbeat", {"message": "keep-alive"})
                        
            finally:
                # 取消订阅
                await council_manager.unsubscribe(meeting_id, queue)
                
        except Exception as e:
            logger.error(f"会议流错误: {e}", exc_info=True)
            yield format_sse("error", {"error": str(e)})
    
    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no"
        }
    )


@app.delete("/api/meetings/{meeting_id}")
async def cancel_meeting(meeting_id: str):
    """
    取消会议
    
    Args:
        meeting_id: 会议ID
        
    Returns:
        取消结果
    """
    try:
        await council_manager.cancel_meeting(meeting_id)
        
        logger.info(f"会议已取消: {meeting_id}")
        
        return {
            "message": "会议已取消",
            "meeting_id": meeting_id
        }
        
    except Exception as e:
        logger.error(f"取消会议错误: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Internal server error")


@app.get("/api/conversations/{conv_id}/meetings")
async def list_conversation_meetings(conv_id: str):
    """
    列出对话的所有会议
    
    Args:
        conv_id: 对话ID
        
    Returns:
        会议列表
    """
    try:
        meetings = await council_manager.list_meetings(conv_id=conv_id)
        
        return {
            "conv_id": conv_id,
            "meetings": meetings,
            "total": len(meetings)
        }
        
    except Exception as e:
        logger.error(f"列出会议错误: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Internal server error")


@app.exception_handler(Exception)
async def global_exception_handler(request, exc):
    """全局异常处理器"""
    logger.error(f"未处理的异常: {exc}", exc_info=True)
    return {
        "detail": "Internal server error",
        "timestamp": get_iso_timestamp()
    }


@app.get("/")
async def root():
    """根路径"""
    return {
        "name": "LLM Council Simplified API",
        "version": "1.0.0",
        "docs": "/docs",
        "redoc": "/redoc"
    }


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8007)